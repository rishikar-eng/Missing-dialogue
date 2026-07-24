"""Convert 'KAMEN RIDER CHARACTER LIST & VOICES.xlsx' into backend/data/voice_bank.json.

The sheet has one row per character with two voice cells (normal + Granute form).
Each cell packs several \"voice name\\nvoice id\" blocks — one per LANGUAGE, the
language encoded in the voice NAME by suffix convention (…Hi, __ta, _te, _MR,
_ML, _BN, _KN, 'Hindi/Tamil', 'TAMIL', …). ElevenLabs IDs are 20-char [A-Za-z0-9].

Run when the spreadsheet changes:
    .venv\\Scripts\\python -m backend.tools.build_voice_bank "KAMEN RIDER CHARACTER LIST & VOICES.xlsx"
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

OUT = Path(__file__).resolve().parents[1] / "data" / "voice_bank.json"

_ID_RE = re.compile(r"^[A-Za-z0-9]{20,22}$")
# Language words as they appear in voice-name tokens (underscore/dash separated).
_LANG_TOKENS = {
    "hi": {"hi", "hindi"}, "ta": {"ta", "tamil"}, "te": {"te", "telugu"},
    "ml": {"ml", "malayalam"}, "mr": {"mr", "marathi"}, "bn": {"bn", "bengali"},
    "kn": {"kn", "kannada"},
}
LANG_LABEL = {"hi": "Hindi", "ta": "Tamil", "te": "Telugu", "ml": "Malayalam",
              "mr": "Marathi", "bn": "Bengali", "kn": "Kannada", "?": "Unknown"}


def _langs_of(voice_name: str) -> list[str]:
    """Language(s) a voice name encodes. 'Hindi/Tamil' voices serve both; for
    compound suffixes like _HI_ta the LAST language token is the actual one.
    Token-based (split on non-letters) because names join tokens with underscores,
    which defeat regex word boundaries (\\b doesn't fire inside `_TAMIL`)."""
    if re.search(r"(?i)hindi\s*/\s*tamil", voice_name):
        return ["hi", "ta"]
    tokens = [t.lower() for t in re.split(r"[^A-Za-z]+", voice_name) if t]
    last: str | None = None
    for t in tokens:
        for lang, words in _LANG_TOKENS.items():
            if t in words:
                last = lang
    return [last or "?"]


def _parse_cell(cell: str, form: str) -> list[dict]:
    """Split a voice cell into [{lang, name, id, form}] entries."""
    out: list[dict] = []
    lines = [ln.strip() for ln in cell.splitlines()]
    block: list[str] = []

    def flush() -> None:
        if not block:
            return
        # ID = last line that looks like an ElevenLabs id; name = the rest.
        idx = next((i for i in range(len(block) - 1, -1, -1) if _ID_RE.match(block[i])), None)
        vid = block[idx] if idx is not None else None
        name = " ".join(b for i, b in enumerate(block) if i != idx).strip(" -")
        if name or vid:
            for lang in _langs_of(name):
                out.append({"lang": lang, "name": name, "id": vid, "form": form})
        block.clear()

    for ln in lines:
        if not ln:
            flush()
        else:
            block.append(ln)
            if _ID_RE.match(ln):  # an ID always ends its block
                flush()
    flush()
    return out


def parse(xlsx: str) -> list[dict]:
    """Parse the character-list workbook -> [{character, raw_name, voices:[{lang,name,id,form}]}].
    Pure (no file writes) so it can be called at run time to build the bank from a freshly
    fetched Box copy — see backend/voices.refresh_from_box."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    bank: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, _img, vid, vid_g = (tuple(row) + (None,) * 4)[:4]
        if not name or not str(name).strip():
            continue
        raw = str(name).strip()
        # Display name = first line; parentheticals/EP notes are context, not identity.
        clean = re.sub(r"\(.*?\)", " ", raw.splitlines()[0]).strip()
        voices: list[dict] = []
        if vid:
            voices += _parse_cell(str(vid), "normal")
        if vid_g:
            voices += _parse_cell(str(vid_g), "granute")
        # Keep rows with NO voice id too (voices == []): the sheet LISTS the character but
        # hasn't recorded a voice yet (e.g. 'Yoshida ma'am'). Carrying them lets the report
        # say "listed, no voice id" instead of the wronger "not in list". A no-id row can
        # never produce a false OK (it has no id to show), so this only sharpens the check.
        bank.append({"character": clean, "raw_name": raw, "voices": voices})
    return bank


def main(xlsx: str) -> int:
    bank = parse(xlsx)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(bank, indent=1, ensure_ascii=False), encoding="utf-8")
    n_ids = sum(1 for e in bank for v in e["voices"] if v["id"])
    print(f"{OUT}: {len(bank)} characters, {sum(len(e['voices']) for e in bank)} voice entries ({n_ids} with IDs)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1] if len(sys.argv) > 1 else "KAMEN RIDER CHARACTER LIST & VOICES.xlsx"))
