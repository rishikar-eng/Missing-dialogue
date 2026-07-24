"""Build the per-episode QC workbook: one .xlsx, one sheet per dub language.

Shape (what the studio asked for):
    Run info   - which files were actually used, settings, parse coverage
    Summary    - one row per language: the at-a-glance comparison + cross-language flags
    <Language> - one sheet per language, each with:
                   Characters  : name, id, mapping + confidence, voice, levels
                   Findings    : missing/misaligned/extra with timestamps, file, confidence
                   Loudness    : too quiet / too hot lines
                   Checks      : track<->character verification + sync warnings

NOTHING here is hardcoded: every analytical cell comes from an /api/analyze run of that
language's tracks. The only reference data is the ElevenLabs VOICE ID, joined per character
from the studio's own 'KAMEN RIDER CHARACTER LIST & VOICES' sheet (see
backend/tools/build_voice_bank.py). A character absent from that sheet gets a blank cell.

(Character pictures were built and then dropped: that sheet only carries portraits for a
minority of the characters that actually appear in an episode - 3 of 11 on the test data -
so the column was mostly empty. `voice_bank_name` on the character records which bank row
was fuzzily matched, which is the useful half of that work and is surfaced below.)
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import voices as voicebank   # aliased: a local 'voices' var exists in _language_sheet

# --- house style -------------------------------------------------------------
_HDR_FILL = PatternFill("solid", fgColor="0D3B66")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=13, color="10233A")
_SECTION_FONT = Font(bold=True, size=11, color="0D3B66")
_MUTED = Font(color="7A8794", size=9)
_THIN = Side(style="thin", color="DBE5EF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
# severity tints: missing = red, misaligned = amber, extra = blue, ok = green
_FILL = {
    "MISSING": PatternFill("solid", fgColor="FBE3E1"),
    "MISMATCH": PatternFill("solid", fgColor="E9D8FD"),   # light purple: delivered, wrong speaker
    "MISALIGNED": PatternFill("solid", fgColor="FDF0D5"),
    "EXTRA": PatternFill("solid", fgColor="E2E8F4"),
    "OK": PatternFill("solid", fgColor="DCEFE0"),
    "WARN": PatternFill("solid", fgColor="FFF3CD"),
}


def _hhmmss(s: float | None) -> str:
    if s is None:
        return ""
    s = max(0.0, float(s))
    h, m = int(s // 3600), int((s % 3600) // 60)
    return f"{h:02d}:{m:02d}:{s % 60:04.1f}"


def _head(ws: Worksheet, row: int, headers: list[str], widths: list[int]) -> int:
    """Write a styled header row and widen its columns.

    A language sheet stacks four tables (characters / findings / loudness / checks) that
    mean different things in the same columns. Assigning width outright made the LAST
    table win for the whole sheet — the characters table ended up wearing the checks
    table's widths and its 30-char track names were squashed into 9. So each column takes
    the WIDEST requirement of any table using it: some columns are roomier than one table
    needs, but nothing is ever truncated.
    """
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font, c.border = _HDR_FILL, _HDR_FONT, _BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        letter = get_column_letter(i)
        cur = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(cur, w)
    return row + 1


def _section(ws: Worksheet, row: int, title: str, note: str = "") -> int:
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    if note:
        ws.cell(row=row, column=3, value=note).font = _MUTED
    return row + 1


# --- per-language sheet ------------------------------------------------------

# Sheet language name -> the voice bank's 2-letter code, so each sheet shows THAT
# language's ElevenLabs voice. Without this, every sheet showed whichever voice was first
# in the bank (Hindi), i.e. a Bengali report carried the Hindi voice id.
_LANG_CODE = {"hindi": "hi", "tamil": "ta", "telugu": "te", "malayalam": "ml",
              "marathi": "mr", "bengali": "bn", "kannada": "kn"}

# A bank match this loose is worth a manual look ('Shoma'->'SHOUMA' ~0.91 is fine; a
# 0.8x scrape onto a similarly-spelled but different character is not).
_VOICE_WEAK = 0.86


def _shares_token(bank_name: str | None, *others: str | None) -> bool:
    """True if the bank name shares a real word (>=3 chars) with any other string. Used to
    clear benign fuzzy matches ('Suga'->'Kenzo Suga', 'Hanto'->'HANTO KARAKEDE') that only
    scored low because the bank spells the name longer — they still name the same person."""
    bt = {t for t in re.split(r"[^a-z0-9]+", (bank_name or "").lower()) if len(t) >= 3}
    return any(bt & {t for t in re.split(r"[^a-z0-9]+", (o or "").lower()) if len(t) >= 3}
               for o in others)


def _voice_check(c: dict[str, Any], lang_code: str | None, lang_label: str,
                 dup_ids: dict[str, list[str]]) -> tuple[str, str, str | None]:
    """Validate one character's assigned ElevenLabs voice id against the studio voice
    list (KAMEN RIDER CHARACTER LIST & VOICES). Returns (status, note, fill_key). Only
    speaking characters are 'audios' worth checking. Statuses:
      OK            - matched a list character and carries this language's id
      not in list   - a NAMED character that isn't in the voice list (studio should add it)
      —             - a generic/bit part (Man, Girl, Crowd) — no dedicated voice expected
      no <lang> id  - matched, but the list has no id for THIS language
      duplicate id  - the id is shared with another character (a list copy-paste error)
      verify match  - a loose match with no shared name word (confirm it's the right voice)
    """
    if not (c.get("line_count") or 0):
        return "", "", None
    name = c.get("name") or ""
    matched = c.get("voice_bank_name")
    if not matched:
        if voicebank._is_generic(name):
            return "—", "generic / bit part — no dedicated voice expected", None
        return "not in list", "no matching character in the studio voice list", "MISSING"
    all_voices = c.get("voices") or []
    lv = [x for x in all_voices if x.get("lang") == lang_code and x.get("id")]
    v = next((x for x in lv if x.get("form") == "normal"), lv[0] if lv else None)
    if not v:
        # A generic bit-part (Kid/Crowd) that merely matched a listed-but-unvoiced row is
        # still "no dedicated voice expected" — don't flag it; only NAMED roles matter here.
        if voicebank._is_generic(name):
            return "—", "generic / bit part — no dedicated voice expected", None
        if not any(x.get("id") for x in all_voices):
            return "no voice id", f"'{matched}' is listed but has no voice id recorded", "WARN"
        return f"no {lang_label} id", f"'{matched}' has no {lang_label} voice id in the list", "WARN"
    shared = [s for s in dup_ids.get(v["id"], []) if s != matched]
    if shared:
        return "duplicate id", "same voice id as: " + ", ".join(shared), "MISMATCH"
    score = c.get("voice_match_score")
    if score is not None and score < _VOICE_WEAK and not _shares_token(matched, name, c.get("channel")):
        return "verify match", f"matched to '{matched}' only loosely ({score:.0%}) — confirm the voice", "WARN"
    return "OK", "", "OK"


def _language_sheet(wb: Workbook, lang: str, res: dict[str, Any]) -> None:
    ws = wb.create_sheet(title=lang[:31])
    chars = res.get("characters") or []
    align = res.get("alignment") or {}
    errors = align.get("errors") or []
    summary = align.get("summary") or {}

    # Voice-ID validation against the studio voice list, per language — computed once and
    # shared by the delivery-health summary and the per-character table below.
    lang_code = _LANG_CODE.get(lang.lower())
    dup_ids = voicebank.duplicate_voice_ids()
    vlist = [_voice_check(c, lang_code, lang.title(), dup_ids) for c in chars]

    ws.cell(row=1, column=1, value=f"{lang} — dialogue QC").font = _TITLE_FONT
    ws.cell(row=1, column=4, value=(
        f"{summary.get('n_missing', 0)} missing · {summary.get('n_misaligned', 0)} misaligned · "
        f"{summary.get('n_extra', 0)} extra · tolerance {align.get('tol_s', '?')}s"
    )).font = _MUTED
    r = 3

    # ---- delivery health ----
    # A packaging summary BEFORE the findings: when the stems themselves are the problem
    # (unclaimed tracks, characters with no stem, ambiguous names, repaired mappings), a
    # big missing count is a delivery issue, not N recording gaps — say so up front so
    # the studio fixes the right thing.
    issues_all = res.get("naming_issues") or []
    channels_all = [str(c) for c in (res.get("channels") or [])]
    owned_tracks = {c.get("channel") for c in chars if c.get("channel")} | {
        x for c in chars for x in (c.get("extra_channels") or [])}
    unclaimed_tracks = [ch for ch in channels_all if ch not in owned_tracks]
    no_track = [c for c in chars
                if not c.get("channel") and not c.get("grouped_in") and (c.get("line_count") or 0) > 0]
    n_twin = sum(1 for i in issues_all if i.get("kind") == "twin_merged")
    n_swap = sum(1 for i in issues_all if i.get("kind") == "swap_repaired")
    n_ambig = sum(1 for i in issues_all if i.get("kind") == "ambiguous_name")
    n_reass = sum(1 for i in issues_all if i.get("kind") == "reassigned")
    miss_all = [e for e in errors if e.get("type") == "MISSING"]
    n_check = sum(1 for e in miss_all if e.get("possibly_in_channel"))
    health: list[tuple[str, bool]] = []   # (text, is_warning)
    health.append((
        f"Tracks delivered: {len(channels_all)} · mapped: {len(owned_tracks)}"
        + (f" · UNCLAIMED: {', '.join(unclaimed_tracks[:4])}"
           + ("…" if len(unclaimed_tracks) > 4 else "") if unclaimed_tracks else ""),
        bool(unclaimed_tracks)))
    if no_track:
        health.append((
            "Characters with NO stem: "
            + ", ".join(f"{c.get('name')} ({c.get('line_count')} lines)" for c in no_track[:6])
            + ("…" if len(no_track) > 6 else ""), True))
    if n_twin or n_swap or n_reass or n_ambig:
        bits = []
        if n_twin: bits.append(f"{n_twin} split stem{'s' if n_twin != 1 else ''} merged")
        if n_swap: bits.append(f"{n_swap} label swap{'s' if n_swap != 1 else ''} repaired")
        if n_reass: bits.append(f"{n_reass} track{'s' if n_reass != 1 else ''} reassigned by voice")
        if n_ambig: bits.append(f"{n_ambig} ambiguous filename{'s' if n_ambig != 1 else ''}")
        health.append(("Mapping repairs: " + " · ".join(bits) + "  (details under CHECKS)", True))
    n_vni = sum(1 for s, _, _ in vlist if s == "not in list")
    n_vdup = sum(1 for s, _, _ in vlist if s == "duplicate id")
    n_vno = sum(1 for s, _, _ in vlist if s.startswith("no "))
    n_vweak = sum(1 for s, _, _ in vlist if s == "verify match")
    if n_vni or n_vdup or n_vno or n_vweak:
        bits = []
        if n_vni: bits.append(f"{n_vni} not in the voice list")
        if n_vdup: bits.append(f"{n_vdup} DUPLICATE voice id")
        if n_vno: bits.append(f"{n_vno} missing this language's id")
        if n_vweak: bits.append(f"{n_vweak} weak name match")
        health.append(("Voice-ID check: " + " · ".join(bits)
                       + "  (see the 'Voice ID check' column)", True))
    if miss_all:
        health.append((
            f"Missing lines: {len(miss_all)} — {len(miss_all) - n_check} confirmed"
            + (f" · {n_check} CHECK-DELIVERY (speech found on an unclaimed track at that slot)"
               if n_check else ""), n_check > 0))
    if any(w for _, w in health) or unclaimed_tracks:
        r = _section(ws, r, "DELIVERY HEALTH", "packaging state of this language's stems")
        for text, warn in health:
            cell = ws.cell(row=r, column=1, value=text)
            cell.alignment = Alignment(vertical="center")
            if warn:
                ws.cell(row=r, column=1).fill = _FILL["WARN"]
            r += 1
        r += 1

    # ---- characters ----
    r = _section(ws, r, "CHARACTERS", "mapping confidence is voice-timeline agreement, not a guess")
    hdr = ["Character", "ID", "Lines", "Dialogue (s)", "Mapped track (file)",
           "Mapped by", "Confidence", "Delivered", "Voice matched as", "ElevenLabs voice",
           "Voice ID", "Voice ID check", "Level min…max (dBFS)", "Reviewer verdict"]
    widths = [22, 16, 7, 12, 30, 11, 11, 11, 20, 22, 24, 15, 18, 18]
    r = _head(ws, r, hdr, widths)
    char_start = r

    # per-character delivered% = fraction of the character's lines actually on THEIR own
    # track. A line that's MISSING (nobody said it) OR MISMATCH (another speaker said it)
    # is not on this character's track, so both count against their delivery rate; the
    # separate Mismatch count shows how many went to the wrong speaker.
    miss_by_char: dict[str, int] = {}
    for e in errors:
        if e.get("type") in ("MISSING", "MISMATCH") and e.get("character"):
            miss_by_char[e["character"]] = miss_by_char.get(e["character"], 0) + 1

    # Mapping confidence comes from content_map's voice-timeline verification, which
    # reports it per naming-issue (precision = "is this track really that character?").
    # A plain NAME match carries no score — we leave it blank rather than invent one.
    conf_by_char: dict[str, float] = {}
    for it in (res.get("naming_issues") or []):
        cid, prec = it.get("character"), it.get("precision")
        if cid and prec is not None:
            conf_by_char[cid] = prec

    for c, vchk in zip(chars, vlist):
        lines = c.get("line_count") or 0
        mapped = bool(c.get("channel")) or bool(c.get("grouped_in"))
        missed = miss_by_char.get(c.get("id"), 0)
        # A character with NO track delivers nothing — but it produces no per-line
        # MISSING findings either (it's reported as unmapped/"no audio"), so a naive
        # 1 - missed/lines would score it 100% delivered. That inversion is the worst
        # possible error here: it would tell the studio an undelivered character is fine.
        delivered = None if not lines else (0.0 if not mapped else 1 - missed / lines)
        voices = c.get("voices") or []
        # THIS language's voice only — never fall back to another language's id, or a
        # Bengali sheet would show the Hindi voice. Prefer the normal form; blank when the
        # bank has no entry for this language (we show nothing rather than a wrong id).
        lang_voices = [x for x in voices if x.get("lang") == lang_code and x.get("id")]
        v = next((x for x in lang_voices if x.get("form") == "normal"),
                 lang_voices[0] if lang_voices else None)

        row_vals = [
            c.get("name"),
            c.get("id"),
            lines,
            round(c.get("total_speech_s") or 0, 1),
            # Twin/pickup stems merged into this character (split deliveries) are listed
            # after the primary — their lines are checked as a UNION of all these stems.
            ((c.get("channel") or "") + "".join(f"  + {x}" for x in (c.get("extra_channels") or []))
             if c.get("channel")
             else ("↳ in " + c["grouped_in"] if c.get("grouped_in") else "— no audio —")),
            c.get("mapped_by") or "",
            None,                                   # confidence (set below, needs format)
            None,                                   # delivered (set below, needs format)
            # WHICH voice-bank row was fuzzily matched ('Shoma' -> 'SHOUMA'). Shown so a
            # wrong voice is visible as a wrong match, not just a wrong-looking ID.
            c.get("voice_bank_name") or "",
            (v or {}).get("name", ""),
            (v or {}).get("id", ""),
            vchk[0],                                # voice-ID check status (col 12)
            (f"{c['level_min_dbfs']:.0f} … {c['level_max_dbfs']:.0f}"
             if c.get("level_min_dbfs") is not None and c.get("level_max_dbfs") is not None else ""),
            "",                                     # reviewer verdict — left blank on purpose
        ]
        for i, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(i == 5))
        # colour the voice-ID check cell (green ok / red not-in-list / purple dup / amber
        # weak|no-id) and attach the detail as a hover comment.
        vstatus, vnote, vfill = vchk
        if vstatus and vfill:
            vc = ws.cell(row=r, column=12)
            vc.fill = _FILL[vfill]
            if vnote:
                vc.comment = Comment(vnote, "QC")
        # confidence: only content-verified mappings have a real number (see above)
        cc = ws.cell(row=r, column=7, value=conf_by_char.get(c.get("id")))
        cc.number_format = "0%"
        dc = ws.cell(row=r, column=8, value=delivered)
        dc.number_format = "0%"
        # colour the mapping cell: green = mapped, amber = grouped, red = nothing
        mc = ws.cell(row=r, column=5)
        mc.fill = _FILL["OK"] if c.get("channel") else (_FILL["WARN"] if c.get("grouped_in") else _FILL["MISSING"])
        r += 1

    # NB: no autofilter here. Excel/openpyxl allow exactly ONE per worksheet, and the
    # Findings block below claims it — that's the table worth filtering (90-190 rows vs
    # ~20 characters). Setting it here too just got silently overwritten.
    r += 1

    # ---- findings ----
    r = _section(ws, r, "FINDINGS", "timestamps are the script/original timeline; MISMATCH = another speaker delivered the line; EXTRA is the dub file's own")
    hdr2 = ["#", "Type", "Character", "Start", "End", "Script line", "Start (s)", "End (s)",
            "Coverage", "Drift (s)", "Track (file)", "Delivered by", "Severity", "Reviewer verdict"]
    r = _head(ws, r, hdr2, [5, 12, 20, 10, 10, 46, 10, 10, 10, 9, 28, 22, 10, 18])
    f_start = r
    # Group by type first — MISSING, MISMATCH, MISALIGNED, EXTRA — and by timeline within
    # each type. Reviewers work missing lines first (the real gaps), so they lead the table.
    _type_rank = {"MISSING": 0, "MISMATCH": 1, "MISALIGNED": 2, "EXTRA": 3}
    by_start = sorted(errors, key=lambda e: (
        _type_rank.get(e.get("type"), 4),
        e.get("script_start_s") if e.get("script_start_s") is not None else e.get("audio_start_s") or 0))
    for n, e in enumerate(by_start, start=1):
        t = e.get("type")
        st = e.get("script_start_s") if e.get("script_start_s") is not None else e.get("audio_start_s")
        en = e.get("script_end_s") if e.get("script_end_s") is not None else e.get("audio_end_s")
        name = next((c.get("name") for c in chars if c.get("id") == e.get("character")), e.get("character") or "")
        deliv = e.get("delivered_by_channel") or ""
        if deliv and e.get("delivered_by_character"):
            by_name = next((c.get("name") for c in chars if c.get("id") == e.get("delivered_by_character")),
                           e.get("delivered_by_character"))
            deliv += f"  ({by_name})"
        # MISSING tier: speech on an UNCLAIMED track at this slot → the line may exist
        # under a bad label. Point at the track; the row is amber (below), not red.
        if t == "MISSING" and e.get("possibly_in_channel"):
            deliv = f"? check '{e['possibly_in_channel']}'"
        vals = [n, t, name, _hhmmss(st), _hhmmss(en),
                e.get("text") or e.get("message") or "",
                round(st, 3) if st is not None else None, round(en, 3) if en is not None else None,
                None,                                        # Coverage (col 9, set below)
                round(e["drift_s"], 2) if e.get("drift_s") is not None else None,
                e.get("channel") or "", deliv, e.get("severity") or "", ""]
        for i, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(i == 6))
        cov = ws.cell(row=r, column=9, value=e.get("coverage"))
        cov.number_format = "0%"
        # check-delivery missing rows render amber (packaging suspicion), confirmed red
        if t == "MISSING" and e.get("possibly_in_channel"):
            ws.cell(row=r, column=2).fill = _FILL["WARN"]
        else:
            ws.cell(row=r, column=2).fill = _FILL.get(t, _FILL["WARN"])
        r += 1
    if r > f_start:
        ws.auto_filter.ref = f"A{f_start - 1}:{get_column_letter(len(hdr2))}{r - 1}"
    else:
        ws.cell(row=r, column=1, value="No findings.").font = _MUTED
        r += 1
    r += 1

    # ---- loudness ----
    loud = res.get("loudness_flags") or []
    if loud:
        r = _section(ws, r, "LOUDNESS", "delivered lines that are too quiet or near clipping")
        r = _head(ws, r, ["Type", "Character", "Start", "Script line", "Level (dBFS)",
                          "Peak (dBFS)", "Track (file)", "Detail"],
                  [11, 20, 10, 44, 13, 12, 28, 52])
        for x in loud:
            vals = [x.get("type"), x.get("character"), _hhmmss(x.get("script_start_s")), x.get("text"),
                    round(x["level_dbfs"], 1) if x.get("level_dbfs") is not None else None,
                    round(x["peak_dbfs"], 1) if x.get("peak_dbfs") is not None else None,
                    x.get("channel"), x.get("message")]
            for i, val in enumerate(vals, start=1):
                c2 = ws.cell(row=r, column=i, value=val)
                c2.border = _BORDER
                c2.alignment = Alignment(vertical="top", wrap_text=(i in (4, 8)))
            ws.cell(row=r, column=1).fill = _FILL["WARN"]
            r += 1
        r += 1

    # ---- checks (mapping + sync) ----
    issues = res.get("naming_issues") or []
    syncs = align.get("sync_warnings") or []
    if issues or syncs:
        r = _section(ws, r, "TRACK ↔ CHARACTER CHECKS", "verify by listening in the app")
        r = _head(ws, r, ["Kind", "Character", "Track (file)", "Recall", "Precision", "Detail"],
                  [18, 20, 28, 9, 10, 72])
        for it in issues:
            vals = [it.get("kind"), it.get("character_name") or it.get("labelled_character_name") or "",
                    it.get("channel") or "", None, None, it.get("message")]
            for i, val in enumerate(vals, start=1):
                c3 = ws.cell(row=r, column=i, value=val)
                c3.border = _BORDER
                c3.alignment = Alignment(vertical="top", wrap_text=(i == 6))
            for col, k2 in ((4, "recall"), (5, "precision")):
                cc2 = ws.cell(row=r, column=col, value=it.get(k2))
                cc2.number_format = "0%"
            r += 1
        for w in syncs:
            for i, val in enumerate(["sync", w.get("character") or "", w.get("channel") or "",
                                     None, None, w.get("message")], start=1):
                c4 = ws.cell(row=r, column=i, value=val)
                c4.border = _BORDER
                c4.alignment = Alignment(vertical="top", wrap_text=(i == 6))
            ws.cell(row=r, column=1).fill = _FILL["WARN"]
            r += 1

    ws.freeze_panes = "A4"


# --- run-info + summary ------------------------------------------------------

def _run_info(wb: Workbook, meta: dict[str, Any], per_lang: dict[str, dict[str, Any]]) -> None:
    ws = wb.create_sheet("Run info", 0)
    ws.cell(row=1, column=1, value="Dialogue QC — run info").font = _TITLE_FONT
    r = 3
    for label, val in [
        ("Episode", meta.get("episode", "")),
        ("Generated", meta.get("generated_at", "")),
        ("Script file", meta.get("script_path", "")),
        ("Original audio", meta.get("original_audio_path", "") or "— none —"),
        ("Tolerance (s)", meta.get("tol_s", "")),
        ("Languages analysed", ", ".join(per_lang)),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(val))
        r += 1
    r += 1

    r = _section(ws, r, "SOURCE FILES USED", "exactly what this report was computed from")
    r = _head(ws, r, ["Language", "Tracks folder", "Tracks", "Script lines parsed", "Parse warnings"],
              [16, 62, 8, 18, 54])
    for lang, res in per_lang.items():
        ps = res.get("parse_stats") or {}
        warn = ""
        if ps and ps.get("dropped"):
            warn = (f"⚠ {ps['dropped']} of {ps['candidates']} dialogue rows could NOT be parsed "
                    f"— those lines were NOT checked")
        vals = [lang, res.get("_audio_dir", ""), len(res.get("channels") or []),
                f"{ps.get('parsed', res.get('n_segments', ''))}" + (f" of {ps['candidates']}" if ps else ""),
                warn]
        for i, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(i in (2, 5)))
        if warn:
            ws.cell(row=r, column=5).fill = _FILL["MISSING"]
        r += 1
    ws.column_dimensions["A"].width = 18
    ws.freeze_panes = "A2"


def _summary(wb: Workbook, per_lang: dict[str, dict[str, Any]]) -> None:
    """The sheet the boss opens first: all languages side by side, plus the
    cross-language signal that only exists because we ran all 6 together."""
    ws = wb.create_sheet("Summary", 0)
    ws.cell(row=1, column=1, value="Summary — all languages").font = _TITLE_FONT
    r = 3
    r = _head(ws, r, ["Language", "Tracks", "Characters", "Missing", "Mismatch", "Misaligned",
                      "Extra", "No audio", "Loudness", "Sync warnings"],
              [16, 8, 12, 10, 10, 12, 9, 10, 10, 14])
    start = r
    for lang, res in per_lang.items():
        s = (res.get("alignment") or {}).get("summary") or {}
        chars = res.get("characters") or []
        no_audio = sum(1 for c in chars if not c.get("channel") and not c.get("grouped_in") and (c.get("line_count") or 0) > 0)
        vals = [lang, len(res.get("channels") or []), len(chars), s.get("n_missing", 0),
                s.get("n_mismatch", 0), s.get("n_misaligned", 0), s.get("n_extra", 0), no_audio,
                len(res.get("loudness_flags") or []),
                len((res.get("alignment") or {}).get("sync_warnings") or [])]
        for i, val in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = _BORDER
        for col, tone in ((4, "MISSING"), (5, "MISMATCH"), (6, "MISALIGNED"), (8, "MISSING")):
            if (ws.cell(row=r, column=col).value or 0) > 0:
                ws.cell(row=r, column=col).fill = _FILL[tone]
        r += 1
    ws.auto_filter.ref = f"A{start - 1}:J{r - 1}"
    r += 2

    # Cross-language consistency — the whole point of one workbook per episode.
    #
    # Judged on the missing RATE per character, never on "has >=1 missing finding": a lead
    # with 78 lines and one dropped line appears in every language's missing list, and
    # calling that "missing everywhere -> script/mapping issue" sends the studio chasing a
    # phantom. Only a character who is *substantially* absent in EVERY language is
    # evidence of a script/mapping problem rather than six dub teams failing identically.
    langs = list(per_lang)
    r = _section(ws, r, "CROSS-LANGUAGE CHECK",
                 "same character, every language — where a gap repeats, the script/mapping is "
                 "the more likely cause than the dub")
    r = _head(ws, r, ["Character", "Missing lines per language", "Languages affected",
                      "Worst miss rate", "Reading"], [22, 40, 16, 14, 60])

    MOSTLY_ABSENT = 0.5     # >= half a character's lines missing = they're effectively absent
    stats: dict[str, dict[str, tuple[int, int]]] = {}   # name -> lang -> (missed, lines)
    for lang, res in per_lang.items():
        chars = {c.get("id"): c for c in (res.get("characters") or [])}
        missed: dict[str, int] = {}
        for e in ((res.get("alignment") or {}).get("errors") or []):
            if e.get("type") == "MISSING" and e.get("character"):
                missed[e["character"]] = missed.get(e["character"], 0) + 1
        for cid, c in chars.items():
            n = missed.get(cid, 0)
            lines = c.get("line_count") or 0
            # a character with no track at all is absent even without per-line findings
            if not (c.get("channel") or c.get("grouped_in")) and lines:
                n = lines
            if n:
                stats.setdefault(c.get("name") or cid, {})[lang] = (n, lines)

    rows = []
    for name, per in stats.items():
        rated = {lg: m / l for lg, (m, l) in per.items() if l}      # language -> miss rate
        worst = max(rated.values()) if rated else 0.0
        absent = [lg for lg in langs if rated.get(lg, 0) >= MOSTLY_ABSENT]  # ~fully missing
        partial = [lg for lg in langs if 0 < rated.get(lg, 0) < MOSTLY_ABSENT]  # a few lines
        delivered = [lg for lg in langs if lg not in per]           # no missing at all
        n = len(langs)
        if len(absent) == n:
            reading, tone = ("Absent in EVERY language — look at the script/mapping first, "
                             "not the dub", "WARN")
        elif absent:
            # Fully missing in SOME languages but not others: the languages that DID deliver
            # prove the line exists, so the fully-missing ones are a delivery/mapping gap
            # there — not a script problem. Name them explicitly (the imprecise old wording
            # called this "a few lines drop in every language").
            bits = [f"fully missing in {', '.join(absent)}"]
            if partial:
                bits.append(f"a few lines in {', '.join(partial)}")
            if delivered:
                bits.append(f"delivered in {', '.join(delivered)}")
            reading = "; ".join(bits) + f" — check delivery/mapping for {', '.join(absent)}"
            tone = "MISSING"
        elif len(per) == n and n > 1:
            reading, tone = ("A few lines drop in every language — usually the same hard lines; "
                             "check those timings in the script", "WARN")
        else:
            reading, tone = (f"Gap in {', '.join(per)} only — the other languages delivered these "
                             f"lines, so it looks like a real dub gap", "MISSING")
        rows.append((name, per, worst, reading, tone))

    for name, per, worst, reading, tone in sorted(rows, key=lambda x: (-len(x[1]), -x[2], x[0])):
        detail = ", ".join(f"{lg} {m}/{l}" for lg, (m, l) in per.items())
        for i, val in enumerate([name, detail, len(per), None, reading], start=1):
            c = ws.cell(row=r, column=i, value=val)
            c.border = _BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(i in (2, 5)))
        wc = ws.cell(row=r, column=4, value=worst)
        wc.number_format = "0%"
        ws.cell(row=r, column=5).fill = _FILL[tone]
        r += 1
    if not rows:
        ws.cell(row=r, column=1, value="No missing lines in any language.").font = _MUTED
    ws.freeze_panes = "A4"


def build_workbook(meta: dict[str, Any], per_lang: dict[str, dict[str, Any]], out_path: str | Path) -> Path:
    """meta: {episode, generated_at, script_path, original_audio_path, tol_s}
    per_lang: {"Malayalam": <analyze result dict + '_audio_dir'>, ...} — insertion-ordered."""
    wb = Workbook()
    wb.remove(wb.active)                       # drop the default sheet
    for lang, res in per_lang.items():
        _language_sheet(wb, lang, res)
    # Both insert at 0, so the LAST one inserted ends up first:
    #   -> [Run info, Summary, <languages...>]
    # (inserting Summary at index 1 instead lands it *after* the first language.)
    _summary(wb, per_lang)
    _run_info(wb, meta, per_lang)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
