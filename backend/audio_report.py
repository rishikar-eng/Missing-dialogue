"""Workbook for an audio-only (scriptless) QC run — one sheet, triage-ordered.





The sound team verifies flags in Pro Tools, so rows are ordered the way they should be


checked: MISSING by confidence (high → low) then time, then EXTRA and MISALIGNED. Times are


given both in seconds and min:sec (they scrub in Audacity/Pro Tools, not in seconds).


"""


from __future__ import annotations





from typing import Any








def _ms(t: float | None) -> str:


    if t is None:


        return ""


    return f"{int(t // 60)}:{t % 60:04.1f}"




def _unchecked_why(s: dict[str, Any]) -> str:
    """"26 dub speaks there, 4 not one line" — or "" when the report cannot say.

    audio_qc owns the gate vocabulary but drags in numpy, and a workbook must still build if
    it is unimportable, so this degrades to the bare count rather than failing the run."""
    try:
        from .audio_qc import unchecked_breakdown
        return ", ".join(f"{n} {r}" for r, n in unchecked_breakdown(s).items())
    except Exception:  # noqa: BLE001
        return ""








def build_audio_workbook(report: dict[str, Any], meta: dict[str, Any], out_path: str) -> str:


    from openpyxl import Workbook


    from openpyxl.styles import Alignment, Font, PatternFill





    wb = Workbook()


    ws = wb.active


    ws.title = "Audio QC"





    s = report.get("summary", {})


    conf = s.get("n_missing_by_confidence", {}) or {}


    head = [


        ("Series", meta.get("series", "")),


        ("Episode", meta.get("episode", "")),


        ("Original (reference)", meta.get("original", "")),


        ("Dub checked", meta.get("dub", "")),


        ("Generated", meta.get("generated_at", "")),


        # Describe what was ACTUALLY compared. This line used to read "original full mix vs
        # dub full mix" unconditionally, which is false in every word for POA: a video
        # original against a dialogue-only or speaker-summed dub.
        ("Mode", "audio-only (no script) — %s vs %s"
                 % (meta.get("original_kind") or "original", meta.get("dub_kind") or "dub")),


        ("Original lines found", s.get("n_original_regions", 0)),


        # Says WHY the unverifiable lines were unverifiable. Most are the dub speaking over
        # the slot; calling them all "unreadable" (as this did) blames the delivered audio.
        ("Verifiable coverage", f"{s.get('coverage', 0.0):.0%} "


                                f"({s.get('n_unchecked', 0)} lines not verifiable"
                                + (f" — {_unchecked_why(s)}" if _unchecked_why(s) else "")
                                + ")"),


        ("MISSING flags", f"{s.get('n_missing', 0)}  "


                          f"(high {conf.get('high', 0)} / medium {conf.get('medium', 0)}"


                          f" / low {conf.get('low', 0)})"),


        ("EXTRA flags", s.get("n_extra", 0)),


        ("MISALIGNED flags", s.get("n_misaligned", 0)),


    ]


    bold = Font(bold=True)


    for k, v in head:


        ws.append([k, v])


        ws.cell(row=ws.max_row, column=1).font = bold


    ws.append([])





    cols = ["Verify order", "Type", "Confidence", "Stability", "Start", "End", "Start (s)", "End (s)",


            "Original line (transcribed)", "Best dub match", "Dub under the line", "Part of", "What to check"]


    ws.append(cols)


    hdr_row = ws.max_row


    fill = PatternFill("solid", fgColor="1F4E79")


    for c in range(1, len(cols) + 1):


        cell = ws.cell(row=hdr_row, column=c)


        cell.font = Font(bold=True, color="FFFFFF")


        cell.fill = fill





    # fragments (1-2 word reference lines) rank below every real finding, whatever their tier
    order = {"high": 0, "medium": 1, "low": 2}


    missing = sorted((e for e in report.get("errors", []) if e["type"] == "MISSING"),


                     key=lambda e: (1 if e.get("fragment") else 0,
                                    order.get(e.get("confidence"), 3),


                                    e.get("script_start_s") or 0.0))


    others = sorted((e for e in report.get("errors", []) if e["type"] != "MISSING"),


                    key=lambda e: (e["type"], e.get("script_start_s") or 0.0))


    tint = {"high": "FCE4E4", "medium": "FFF2CC", "low": "EDEDED"}


    for i, e in enumerate(missing + others, start=1):


        st, en = e.get("script_start_s"), e.get("script_end_s")


        if e["type"] == "MISSING":


            # KEY THE NOTE OFF THE MEASURED FILL, not off the row type. Every MISSING row
            # used to read "the original line appears to have no dub" — which the row's own
            # "Dub under the line" column contradicted whenever the fill was speech-like.
            # POA EP31 shipped 4 rows saying "no dub" and "dub speaks here" side by side.
            _f = e.get("fill")
            if _f == "speech-like":
                note = ("The dub SPEAKS under this line — check whether it is this line "
                        "delivered differently, or a neighbouring line bleeding in.")
            elif _f == "ambience":
                note = ("Only ambience/walla under this line — no dialogue, but the dub is "
                        "not silent here.")
            elif _f == "silence":
                note = "The dub is silent under this line — listen for a genuine drop."
            else:
                # fill could not be measured (slot too short, or _dub_fill raised)
                note = ("Listen to the dub here — no usable dub audio was measured under "
                        "this line.")


        elif e["type"] == "EXTRA":


            note = "The dub speaks here but the original does not — added line?"


        else:


            note = f"Line present but shifted by {e.get('drift_s', 0):+.1f}s."


        best = e.get("coverage")


        fill = {"silence": "silent (no dub audio)",
                "ambience": "covered by ambience/walla",
                "speech-like": "dub speaks here"}.get(e.get("fill"), "")
        if fill and e.get("fill_dyn_db") is not None:
            fill += f" ({e['fill_rel_db']:+.0f} dB vs speech, {e['fill_dyn_db']:.0f} dB range)"
        blk = ""
        if e.get("sung"):
            blk = ("sung line (music above the vocal) — songs are not dubbed on this series, "
                   "so this is listed but NOT given a Pro Tools marker")
        if e.get("fragment") and not e.get("block") and not e.get("song_reprise") and not e.get("sung"):
            blk = "fragment (1-2 words) — too short to match reliably; check last"
        if e.get("song_reprise") and not e.get("block") and not e.get("sung"):
            # Defensive: this is the ONLY report builder not wrapped in try/except, so a
            # malformed field here aborts the whole run before markers or reference audio are
            # written. Treat anything that is not the expected dict as a bare flag.
            r = e["song_reprise"] if isinstance(e["song_reprise"], dict) else {}
            where = _ms(r.get("matches_at")) if r.get("matches_at") is not None else "elsewhere"
            blk = f"same lyric sung {where} — untranslated song, not dialogue"
        if e.get("block"):
            b = e["block"]
            blk = (f"block {b['id']}: {_ms(b['start'])}-{_ms(b['end'])} "
                   f"({b['n']} lines, no dub audio throughout)")
        ws.append([i, e["type"], e.get("confidence", ""), e.get("stability", ""),


                   _ms(st), _ms(en), st, en, e.get("text") or "",


                   f"{best:.0%}" if best is not None else "", fill, blk, note])


        color = tint.get(e.get("confidence")) if e["type"] == "MISSING" else None


        if color:


            row_fill = PatternFill("solid", fgColor=color)


            for c in range(1, len(cols) + 1):


                ws.cell(row=ws.max_row, column=c).fill = row_fill





    widths = [11, 12, 11, 10, 8, 8, 9, 9, 60, 13, 34, 34, 52]


    for c, w in enumerate(widths, start=1):


        ws.column_dimensions[ws.cell(row=hdr_row, column=c).column_letter].width = w


        for r in range(hdr_row, ws.max_row + 1):


            ws.cell(row=r, column=c).alignment = Alignment(vertical="top",


                                                           wrap_text=(c in (9, 11, 12, 13)))


    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)


    wb.save(out_path)


    return out_path








def build_marker_midi(report, out_path, start_tc_s=0.0, fps=25.0):
    """Pro Tools-importable marker file: every MISSING flag becomes a named Memory
    Location at its timecode (File > Import > MIDI, markers enabled). A true .ptx is
    Avid-proprietary and encrypted — no open tool writes one — so marker MIDI is the
    interchange the sound team can actually use.

    Written for POST, not music, because that is where imports go wrong:
      * TYPE 1 with a dedicated conductor track — several importers ignore a bare type-0.
      * A named marker track, since unnamed tracks get skipped.
      * An SMPTE offset meta event, so a session that starts at 01:00:00:00 does not put
        every marker an hour out. Pass start_tc_s to match the session start.
      * Tempo pinned to 120 BPM with 480 ppqn, i.e. exactly 960 ticks per second, so the
        tick maths is round and a tempo-following import lands where it should.
    `start_tc_s` shifts every marker EARLIER by the session start, so a marker at 6:01 in
    a session starting at 01:00:00:00 is written at 6:01 past that start.
    """
    import struct

    TPQ = 480                      # at 120 BPM -> 960 ticks per second

    def vlq(n):
        out = [n & 0x7F]
        while n > 0x7F:
            n >>= 7
            out.append((n & 0x7F) | 0x80)
        return bytes(reversed(out))

    def track(events):
        data = b"".join(events) + bytes([0, 255, 47, 0])          # + EOT
        return b"MTrk" + struct.pack(">I", len(data)) + data

    # --- conductor track: tempo + SMPTE start, what a post session needs to place ticks
    fps_code = {24: 0, 25: 1, 29.97: 2, 30: 3}.get(round(float(fps), 2), 1)
    t0 = max(0.0, float(start_tc_s))
    hh, mm = int(t0 // 3600), int((t0 % 3600) // 60)
    ss, ff = int(t0 % 60), int(round((t0 % 1) * fps))
    conductor = [
        bytes([0, 255, 3, 8]) + b"Markers ",                       # track name
        bytes([0, 255, 81, 3, 7, 161, 32]),                        # 120 BPM
        bytes([0, 255, 84, 5, (fps_code << 5) | hh, mm, ss, ff, 0]),  # SMPTE offset
        bytes([0, 255, 88, 4, 4, 2, 24, 8]),                       # 4/4 time signature
    ]

    events, prev = [bytes([0, 255, 3, 12]) + b"QC Markers  "], 0
    # SUNG LINES DO NOT GET A MARKER. On a series whose songs are left untranslated a sung
    # line can never be a real finding, and the engineers' complaint about the first PTX
    # batch was exactly this: hundreds of markers over the themes, all correctly ranked low
    # and all useless to click through. 430 of the flags across 15 POA episodes sat inside
    # song regions. They remain in the WORKBOOK — this only keeps them off the timeline, so
    # nothing is hidden, it just stops being something a human has to dismiss one by one.
    for e in sorted((x for x in report.get("errors", [])
                     if x["type"] == "MISSING"
                     and not (x.get("sung") or x.get("song_reprise"))),
                    key=lambda x: x["script_start_s"]):
        t = float(e["script_start_s"])
        # SHORT NAMES ON PURPOSE. Pro Tools clips a marker label to the pixel gap before the
        # next marker, so a long name is simply unreadable wherever markers are dense — which
        # is exactly where the interesting ones sit. The engineer's screenshots showed names
        # arriving as "121" and "MISI27437" for that reason. The timecode is dropped too: the
        # marker's own position already states it, and the Portuguese line travels in the
        # COMMENT, which Pro Tools shows in the Markers window.
        label = ("SONG" if (e.get("block") or e.get("song_reprise")) else
                 "FRAG" if e.get("fragment") else
                 {"high": "MISS-HI", "medium": "MISS-MED"}.get(e.get("confidence"), "MISS-LOW"))
        name = label.encode("ascii", "replace")[:60]
        tick = int(round(t * 2 * TPQ))
        events.append(vlq(tick - prev) + bytes([255, 6]) + vlq(len(name)) + name)
        prev = tick
    data = (b"MThd" + struct.pack(">IHHH", 6, 1, 2, TPQ)           # TYPE 1, two tracks
            + track(conductor) + track(events))
    with open(out_path, "wb") as f:
        f.write(data)
    return out_path


def build_ref_audio(errors, original_path, out_path, timeline_path=None,
                    pad_s=2.5, gap_s=0.6):
    """MISSING-only reference audio, shared by BOTH QC modes: the original-language audio of
    every genuinely MISSING line. `out_path` = stitched (clips back-to-back);
    `timeline_path` = same clips at their real episode timecodes (silent everywhere else),
    for lining up 1:1 against the dub session. Returns the stitched path or None when there
    is nothing missing. (Moved here from episode_runner so the scriptless runner — which
    must not import the Box-fetch stack — can build the same deliverable.)"""
    import numpy as np
    import soundfile as sf
    wins = sorted(
        (max(0.0, e["script_start_s"] - pad_s),
         (e.get("script_end_s") or e["script_start_s"]) + pad_s)
        for e in errors
        if e.get("type") == "MISSING" and e.get("script_start_s") is not None
    )
    merged = []
    for s0, e0 in wins:
        if merged and s0 <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], e0)
        else:
            merged.append([s0, e0])
    if not merged:
        return None
    with sf.SoundFile(str(original_path)) as f:
        sr, total = f.samplerate, len(f)
        gap = np.zeros(int(gap_s * sr), dtype=np.float32)
        chunks = []
        tl = np.zeros(total, dtype=np.float32) if timeline_path else None
        for s0, e0 in merged:
            i0, i1 = max(0, int(s0 * sr)), min(total, int(e0 * sr))
            if i1 <= i0:
                continue
            f.seek(i0)
            d = f.read(i1 - i0, dtype="float32", always_2d=False)
            if getattr(d, "ndim", 1) > 1:
                d = d.mean(axis=1)
            chunks += [d, gap]
            if tl is not None:
                tl[i0:i0 + len(d)] = d
        out = np.concatenate(chunks) if chunks else np.zeros(0, dtype=np.float32)
    if not len(out):
        return None
    sf.write(str(out_path), out, sr, format="FLAC")
    if tl is not None:
        sf.write(str(timeline_path), tl, sr, format="FLAC")
    return out_path


def build_marker_csv(report, out_path, fps=25.0):
    """Marker list as the CSV the studio's converter actually ingests.

    Pro Tools cannot import a plain text marker list natively, so the route to a real .ptx is
    a converter (editingtools.io / EdiMarker). This is the exact layout the sound team has
    already converted successfully — No.,In,Color,Name,Comment with absolute SMPTE at 25 fps,
    zero-based, because the delivered dubs carry BWF TimeReference 0 and need no session
    offset. An earlier version of this function emitted Reaper's layout instead and was called
    by nothing; matching what demonstrably worked beats shipping a second format for someone
    to debug.

    Same contents as the MIDI: sung lines are left off (songs are not dubbed on this series and
    they were 21 of every 33 flags), and names are short because Pro Tools clips a marker label
    to the pixel gap before the next one.
    """
    import csv as _csv

    def smpte(t):
        t = max(0.0, float(t))
        h, m = int(t // 3600), int((t % 3600) // 60)
        sec, f = int(t % 60), int(round((t - int(t)) * fps))
        if f >= fps:                      # rounding can tip a frame past the second
            f, sec = 0, sec + 1
            if sec == 60:
                sec, m = 0, m + 1
        return "%02d:%02d:%02d:%02d" % (h, m, sec, f)

    def label(e):
        if e.get("song_reprise") or e.get("sung"):
            return "Blue", "SONG"
        if e.get("block"):
            return "Orange", "BLOCK"
        if e.get("fragment"):
            return "Green", "FRAG"
        c = (e.get("confidence") or "low").lower()
        return ({"high": "Red", "medium": "Red"}.get(c, "Orange"),
                {"high": "MISS-HI", "medium": "MISS-MED"}.get(c, "MISS-LOW"))

    rows = sorted((x for x in report.get("errors", [])
                   if x["type"] == "MISSING" and x.get("script_start_s") is not None
                   and not (x.get("sung") or x.get("song_reprise"))),
                  key=lambda x: x["script_start_s"])
    with open(out_path, "w", encoding="utf-8-sig", newline="") as fh:
        w = _csv.writer(fh)
        w.writerow(["No.", "In", "Color", "Name", "Comment"])
        for i, e in enumerate(rows, start=1):
            colour, name = label(e)
            w.writerow([i, smpte(e["script_start_s"]), colour, name,
                        (e.get("text") or "").replace(chr(10), " ")])
    return out_path


def build_ref_audio(errors, original_path, out_path, timeline_path=None,
                    pad_s=2.5, gap_s=0.6):
    """MISSING-only reference audio, shared by BOTH QC modes: the original-language audio of
    every genuinely MISSING line. `out_path` = stitched (clips back-to-back);
    `timeline_path` = same clips at their real episode timecodes (silent everywhere else),
    for lining up 1:1 against the dub session. Returns the stitched path or None when there
    is nothing missing. (Moved here from episode_runner so the scriptless runner — which
    must not import the Box-fetch stack — can build the same deliverable.)"""
    import numpy as np
    import soundfile as sf
    wins = sorted(
        (max(0.0, e["script_start_s"] - pad_s),
         (e.get("script_end_s") or e["script_start_s"]) + pad_s)
        for e in errors
        if e.get("type") == "MISSING" and e.get("script_start_s") is not None
    )
    merged = []
    for s0, e0 in wins:
        if merged and s0 <= merged[-1][1]:
            merged[-1][1] = max(merged[-1][1], e0)
        else:
            merged.append([s0, e0])
    if not merged:
        return None
    with sf.SoundFile(str(original_path)) as f:
        sr, total = f.samplerate, len(f)
        gap = np.zeros(int(gap_s * sr), dtype=np.float32)
        chunks = []
        tl = np.zeros(total, dtype=np.float32) if timeline_path else None
        for s0, e0 in merged:
            i0, i1 = max(0, int(s0 * sr)), min(total, int(e0 * sr))
            if i1 <= i0:
                continue
            f.seek(i0)
            d = f.read(i1 - i0, dtype="float32", always_2d=False)
            if getattr(d, "ndim", 1) > 1:
                d = d.mean(axis=1)
            chunks += [d, gap]
            if tl is not None:
                tl[i0:i0 + len(d)] = d
        out = np.concatenate(chunks) if chunks else np.zeros(0, dtype=np.float32)
    if not len(out):
        return None
    sf.write(str(out_path), out, sr, format="FLAC")
    if tl is not None:
        sf.write(str(timeline_path), tl, sr, format="FLAC")
    return out_path


